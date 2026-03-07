"""
LocalFile MCP - 本地文件操作工具
提供文件上传、解析和数据流创建能力
"""

from typing import Dict, Any, List, Optional
from datetime import datetime
import json
import os
import csv
import io
from pathlib import Path

from src.mcp.service import MCPTool
from src.core.config import CONFIG_DIR


UPLOAD_DIR = CONFIG_DIR / "uploads"
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)


class UploadFileTool:
    """上传文件"""

    def get_name(self) -> str:
        return "pbbi_local_upload_file"

    def get_description(self) -> str:
        return "上传文件到服务器，支持CSV、Excel、JSON等格式。返回文件ID和基本信息。"

    def get_parameters(self) -> Dict[str, Any]:
        return {
            "type": "object",
            "properties": {
                "file_name": {
                    "type": "string",
                    "description": "文件名"
                },
                "file_content": {
                    "type": "string",
                    "description": "文件内容（Base64编码或文本内容）"
                },
                "file_type": {
                    "type": "string",
                    "description": "文件类型：csv, json, txt, excel",
                    "enum": ["csv", "json", "txt", "excel"]
                },
                "encoding": {
                    "type": "string",
                    "description": "文件编码，默认utf-8"
                }
            },
            "required": ["file_name", "file_content", "file_type"]
        }

    def execute(self, params: Dict[str, Any]) -> Dict[str, Any]:
        file_name = params.get("file_name")
        file_content = params.get("file_content")
        file_type = params.get("file_type", "csv").lower()
        encoding = params.get("encoding", "utf-8")

        if not file_name or not file_content:
            return {"success": False, "error": "file_name和file_content是必需的"}

        import base64
        import time

        timestamp = int(time.time() * 1000)
        safe_name = "".join(c if c.isalnum() or c in "._-" else "_" for c in file_name)
        stored_name = f"{timestamp}_{safe_name}"
        file_path = UPLOAD_DIR / stored_name

        try:
            if file_type in ["csv", "json", "txt"]:
                try:
                    decoded_content = base64.b64decode(file_content).decode(encoding)
                except Exception:
                    decoded_content = file_content
                
                with open(file_path, "w", encoding=encoding) as f:
                    f.write(decoded_content)
            else:
                try:
                    decoded_content = base64.b64decode(file_content)
                except Exception:
                    return {"success": False, "error": "Excel文件需要Base64编码"}
                
                with open(file_path, "wb") as f:
                    f.write(decoded_content)

            file_size = file_path.stat().st_size

            return {
                "success": True,
                "data": {
                    "file_id": stored_name,
                    "file_name": file_name,
                    "file_type": file_type,
                    "file_size": file_size,
                    "file_path": str(file_path),
                    "uploaded_at": datetime.now().isoformat()
                }
            }
        except Exception as e:
            return {"success": False, "error": f"文件上传失败: {str(e)}"}


class ParseFileTool:
    """解析文件"""

    def get_name(self) -> str:
        return "pbbi_local_parse_file"

    def get_description(self) -> str:
        return "解析已上传的文件，提取数据和字段信息。支持CSV、JSON、Excel格式。"

    def get_parameters(self) -> Dict[str, Any]:
        return {
            "type": "object",
            "properties": {
                "file_id": {
                    "type": "string",
                    "description": "文件ID（上传时返回的file_id）"
                },
                "file_type": {
                    "type": "string",
                    "description": "文件类型：csv, json, excel",
                    "enum": ["csv", "json", "excel"]
                },
                "sheet_name": {
                    "type": "string",
                    "description": "Excel工作表名称（仅Excel文件需要）"
                },
                "delimiter": {
                    "type": "string",
                    "description": "CSV分隔符，默认逗号"
                },
                "encoding": {
                    "type": "string",
                    "description": "文件编码，默认utf-8"
                },
                "preview_rows": {
                    "type": "integer",
                    "description": "预览行数，默认10"
                }
            },
            "required": ["file_id"]
        }

    def execute(self, params: Dict[str, Any]) -> Dict[str, Any]:
        file_id = params.get("file_id")
        file_type = params.get("file_type")
        sheet_name = params.get("sheet_name")
        delimiter = params.get("delimiter", ",")
        encoding = params.get("encoding", "utf-8")
        preview_rows = params.get("preview_rows", 10)

        if not file_id:
            return {"success": False, "error": "file_id是必需的"}

        file_path = UPLOAD_DIR / file_id
        if not file_path.exists():
            return {"success": False, "error": f"文件不存在: {file_id}"}

        try:
            if file_type is None:
                ext = file_path.suffix.lower()
                if ext == ".csv":
                    file_type = "csv"
                elif ext == ".json":
                    file_type = "json"
                elif ext in [".xlsx", ".xls"]:
                    file_type = "excel"
                else:
                    file_type = "csv"

            if file_type == "csv":
                return self._parse_csv(file_path, delimiter, encoding, preview_rows)
            elif file_type == "json":
                return self._parse_json(file_path, encoding, preview_rows)
            elif file_type == "excel":
                return self._parse_excel(file_path, sheet_name, preview_rows)
            else:
                return {"success": False, "error": f"不支持的文件类型: {file_type}"}

        except Exception as e:
            return {"success": False, "error": f"文件解析失败: {str(e)}"}

    def _parse_csv(self, file_path: Path, delimiter: str, encoding: str, preview_rows: int) -> Dict[str, Any]:
        with open(file_path, "r", encoding=encoding) as f:
            reader = csv.reader(f, delimiter=delimiter)
            headers = next(reader)
            
            rows = []
            for i, row in enumerate(reader):
                if i >= preview_rows:
                    break
                rows.append(dict(zip(headers, row)))

            total_rows = sum(1 for _ in open(file_path, "r", encoding=encoding)) - 1

        fields = []
        for header in headers:
            sample_value = rows[0].get(header, "") if rows else ""
            field_type = self._infer_type(sample_value)
            fields.append({
                "name": header,
                "type": field_type,
                "sample_value": sample_value
            })

        return {
            "success": True,
            "data": {
                "file_type": "csv",
                "fields": fields,
                "preview_data": rows,
                "total_rows": total_rows,
                "total_columns": len(headers)
            }
        }

    def _parse_json(self, file_path: Path, encoding: str, preview_rows: int) -> Dict[str, Any]:
        with open(file_path, "r", encoding=encoding) as f:
            data = json.load(f)

        if isinstance(data, list):
            rows = data[:preview_rows]
            total_rows = len(data)
        elif isinstance(data, dict):
            rows = [data]
            total_rows = 1
        else:
            return {"success": False, "error": "JSON格式不支持"}

        if not rows:
            return {"success": False, "error": "JSON数据为空"}

        fields = []
        first_row = rows[0]
        for key, value in first_row.items():
            field_type = self._infer_type(value)
            fields.append({
                "name": key,
                "type": field_type,
                "sample_value": value
            })

        return {
            "success": True,
            "data": {
                "file_type": "json",
                "fields": fields,
                "preview_data": rows,
                "total_rows": total_rows,
                "total_columns": len(fields)
            }
        }

    def _parse_excel(self, file_path: Path, sheet_name: Optional[str], preview_rows: int) -> Dict[str, Any]:
        try:
            import openpyxl
        except ImportError:
            return {"success": False, "error": "需要安装openpyxl库: pip install openpyxl"}

        wb = openpyxl.load_workbook(file_path, read_only=True)
        
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                return {"success": False, "error": f"工作表不存在: {sheet_name}"}
            ws = wb[sheet_name]
        else:
            ws = wb.active
            sheet_name = ws.title

        rows_iter = ws.iter_rows(values_only=True)
        headers = next(rows_iter)
        headers = [str(h) if h else f"column_{i}" for i, h in enumerate(headers)]

        rows = []
        for i, row in enumerate(rows_iter):
            if i >= preview_rows:
                break
            rows.append(dict(zip(headers, row)))

        total_rows = ws.max_row - 1

        fields = []
        for header in headers:
            sample_value = rows[0].get(header, "") if rows else ""
            field_type = self._infer_type(sample_value)
            fields.append({
                "name": header,
                "type": field_type,
                "sample_value": sample_value
            })

        wb.close()

        return {
            "success": True,
            "data": {
                "file_type": "excel",
                "sheet_name": sheet_name,
                "available_sheets": wb.sheetnames if hasattr(wb, 'sheetnames') else [sheet_name],
                "fields": fields,
                "preview_data": rows,
                "total_rows": total_rows,
                "total_columns": len(headers)
            }
        }

    def _infer_type(self, value: Any) -> str:
        if value is None:
            return "text"
        if isinstance(value, bool):
            return "boolean"
        if isinstance(value, int):
            return "integer"
        if isinstance(value, float):
            return "number"
        if isinstance(value, str):
            try:
                float(value)
                if "." in value:
                    return "number"
                return "integer"
            except ValueError:
                pass
            if value.lower() in ["true", "false", "是", "否"]:
                return "boolean"
        return "text"


class CreateDataFlowFromFilesTool:
    """从文件创建数据流"""

    def get_name(self) -> str:
        return "pbbi_local_create_dataflow"

    def get_description(self) -> str:
        return "从已上传的文件创建数据流，并可选地保存为快照"

    def get_parameters(self) -> Dict[str, Any]:
        return {
            "type": "object",
            "properties": {
                "dataflow_name": {
                    "type": "string",
                    "description": "数据流名称"
                },
                "file_id": {
                    "type": "string",
                    "description": "文件ID"
                },
                "file_type": {
                    "type": "string",
                    "description": "文件类型：csv, json, excel",
                    "enum": ["csv", "json", "excel"]
                },
                "selected_fields": {
                    "type": "array",
                    "items": {"type": "string"},
                    "description": "选择的字段列表，不填则使用所有字段"
                },
                "create_snapshot": {
                    "type": "boolean",
                    "description": "是否同时创建快照，默认true"
                },
                "snapshot_name": {
                    "type": "string",
                    "description": "快照名称（create_snapshot为true时使用）"
                }
            },
            "required": ["dataflow_name", "file_id"]
        }

    def execute(self, params: Dict[str, Any]) -> Dict[str, Any]:
        dataflow_name = params.get("dataflow_name")
        file_id = params.get("file_id")
        file_type = params.get("file_type")
        selected_fields = params.get("selected_fields", [])
        create_snapshot = params.get("create_snapshot", True)
        snapshot_name = params.get("snapshot_name")

        if not dataflow_name or not file_id:
            return {"success": False, "error": "dataflow_name和file_id是必需的"}

        file_path = UPLOAD_DIR / file_id
        if not file_path.exists():
            return {"success": False, "error": f"文件不存在: {file_id}"}

        try:
            parse_tool = ParseFileTool()
            parse_result = parse_tool.execute({
                "file_id": file_id,
                "file_type": file_type
            })

            if not parse_result.get("success"):
                return parse_result

            parsed_data = parse_result.get("data", {})
            all_fields = parsed_data.get("fields", [])
            
            if selected_fields:
                fields = [f for f in all_fields if f["name"] in selected_fields]
            else:
                fields = all_fields

            from src.core.database import SessionLocal
            from src.models.config import DataFlow, FieldType

            db = SessionLocal()
            try:
                dataflow = DataFlow(
                    name=dataflow_name,
                    type="local_file",
                    worksheet_id=file_id
                )
                db.add(dataflow)
                db.flush()

                for field in fields:
                    field_type = FieldType(
                        data_flow_id=dataflow.id,
                        field_id=field["name"],
                        field_name=field["name"],
                        data_type=field.get("type", "text"),
                        is_enabled="true"
                    )
                    db.add(field_type)

                db.commit()
                dataflow_id = dataflow.id
            finally:
                db.close()

            result = {
                "success": True,
                "data": {
                    "dataflow_id": dataflow_id,
                    "dataflow_name": dataflow_name,
                    "file_id": file_id,
                    "fields_count": len(fields),
                    "fields": [f["name"] for f in fields]
                }
            }

            if create_snapshot:
                if not snapshot_name:
                    snapshot_name = f"{dataflow_name}_快照"

                preview_data = parsed_data.get("preview_data", [])
                
                if selected_fields:
                    filtered_data = []
                    for row in preview_data:
                        filtered_row = {k: v for k, v in row.items() if k in selected_fields}
                        filtered_data.append(filtered_row)
                else:
                    filtered_data = preview_data

                from src.mcp.database_mcp import CreateSnapshotTableTool
                import time

                create_tool = CreateSnapshotTableTool()
                snapshot_result = create_tool.execute({
                    "name": snapshot_name,
                    "table_name": f"local_{file_id}_{int(time.time())}",
                    "data_flow_id": dataflow_id,
                    "fields": [{"name": f["name"]} for f in fields],
                    "data": filtered_data
                })

                if snapshot_result.get("success"):
                    result["data"]["snapshot_id"] = snapshot_result.get("data", {}).get("snapshot_id")
                    result["data"]["snapshot_name"] = snapshot_name

            return result

        except Exception as e:
            return {"success": False, "error": f"创建数据流失败: {str(e)}"}


class ListUploadedFilesTool:
    """列出已上传文件"""

    def get_name(self) -> str:
        return "pbbi_local_list_files"

    def get_description(self) -> str:
        return "列出所有已上传的文件"

    def get_parameters(self) -> Dict[str, Any]:
        return {
            "type": "object",
            "properties": {
                "file_type": {
                    "type": "string",
                    "description": "按文件类型筛选：csv, json, excel"
                }
            },
            "required": []
        }

    def execute(self, params: Dict[str, Any]) -> Dict[str, Any]:
        file_type = params.get("file_type")

        files = []
        for file_path in UPLOAD_DIR.iterdir():
            if not file_path.is_file():
                continue

            ext = file_path.suffix.lower()
            detected_type = None
            if ext == ".csv":
                detected_type = "csv"
            elif ext == ".json":
                detected_type = "json"
            elif ext in [".xlsx", ".xls"]:
                detected_type = "excel"

            if file_type and detected_type != file_type:
                continue

            stat = file_path.stat()
            parts = file_path.name.split("_", 1)
            original_name = parts[1] if len(parts) > 1 else file_path.name

            files.append({
                "file_id": file_path.name,
                "file_name": original_name,
                "file_type": detected_type,
                "file_size": stat.st_size,
                "uploaded_at": datetime.fromtimestamp(stat.st_ctime).isoformat()
            })

        files.sort(key=lambda x: x["uploaded_at"], reverse=True)

        return {
            "success": True,
            "data": {
                "files": files,
                "total": len(files)
            }
        }


class DeleteUploadedFileTool:
    """删除已上传文件"""

    def get_name(self) -> str:
        return "pbbi_local_delete_file"

    def get_description(self) -> str:
        return "删除已上传的文件"

    def get_parameters(self) -> Dict[str, Any]:
        return {
            "type": "object",
            "properties": {
                "file_id": {
                    "type": "string",
                    "description": "文件ID"
                }
            },
            "required": ["file_id"]
        }

    def execute(self, params: Dict[str, Any]) -> Dict[str, Any]:
        file_id = params.get("file_id")

        if not file_id:
            return {"success": False, "error": "file_id是必需的"}

        file_path = UPLOAD_DIR / file_id
        if not file_path.exists():
            return {"success": False, "error": f"文件不存在: {file_id}"}

        try:
            file_path.unlink()
            return {
                "success": True,
                "data": {
                    "deleted_file_id": file_id
                }
            }
        except Exception as e:
            return {"success": False, "error": f"删除文件失败: {str(e)}"}


def register_localfile_tools(mcp_service):
    """注册LocalFile MCP工具"""
    tools = [
        UploadFileTool(),
        ParseFileTool(),
        CreateDataFlowFromFilesTool(),
        ListUploadedFilesTool(),
        DeleteUploadedFileTool(),
    ]

    for tool in tools:
        mcp_service.register_tool(MCPTool(
            name=tool.get_name(),
            description=tool.get_description(),
            parameters=tool.get_parameters()
        ))
        mcp_service.register_handler(tool.get_name(), tool.execute)

    return tools
